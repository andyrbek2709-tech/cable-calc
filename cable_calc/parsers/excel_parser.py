"""
Excel-парсер кабельных журналов (.xlsx / .xls / .xlsm).
Поддерживает два режима:
  - Шапка найдена: используем колонки по ключевым словам
  - Шапка не найдена: парсим каждую строку эвристически
"""
import os
import re
from typing import List, Dict, Optional

import openpyxl
from .models import CableJournalRow, ParseResult
from .utils import parse_section, parse_length, parse_cable_mark, parse_voltage, clean_ocr

# Ключевые слова для определения колонок
COL_KEYS = {
    "id":      ["обозначение", "позиция", "марка кабеля", "номер", "cable id", "designation"],
    "name":    ["наименование", "назначение", "name", "description", "от"],
    "from":    ["начало", "откуда", "from", "источник"],
    "to":      ["конец", "куда", "to", "приёмник", "потребитель"],
    "mark":    ["марка", "тип", "brand", "type кабель"],
    "section": ["сечение", "section", "мм", "провод"],
    "length":  ["длина", "length", "м ", "протяжённость"],
    "voltage": ["напряжение", "voltage", "кв", "kv"],
}


def _detect_header(ws) -> Optional[Dict[str, int]]:
    """Ищем строку-заголовок в первых 10 строках. Возвращает {field: col_index}."""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        row_str = [str(c or "").lower() for c in row]
        mapping = {}
        for field, keywords in COL_KEYS.items():
            for idx, cell in enumerate(row_str):
                if any(kw in cell for kw in keywords):
                    mapping[field] = idx
                    break
        if len(mapping) >= 3:  # Нашли хотя бы 3 поля — считаем заголовком
            return mapping
    return None


def _row_from_mapping(cells: list, mapping: Dict[str, int], row_num: int) -> Optional[CableJournalRow]:
    def get(field) -> str:
        idx = mapping.get(field)
        if idx is None or idx >= len(cells):
            return ""
        return str(cells[idx] or "").strip()

    row = CableJournalRow(row_num=row_num)
    row.cable_id = get("id")[:40]
    row.cable_name = get("name")[:80]
    row.from_point = get("from")[:60]
    row.to_point = get("to")[:60]

    mark_cell = get("mark")
    row.cable_mark = parse_cable_mark(mark_cell) or mark_cell[:30]

    phases, s, zs, raw = parse_section(get("section"))
    if raw:
        row.phases, row.section_mm2, row.zero_section_mm2, row.section_str = phases, s, zs, raw

    len_cell = get("length")
    try:
        row.length_m = float(str(len_cell).replace(",", ".")) if len_cell else 0.0
    except ValueError:
        row.length_m = parse_length(len_cell)

    v = get("voltage")
    if v:
        row.voltage_kv = parse_voltage(v)

    row.source_line = " | ".join(str(c or "") for c in cells)

    if not row.cable_mark and not row.section_str and row.length_m == 0:
        return None
    return row


def _row_heuristic(cells: list, row_num: int) -> Optional[CableJournalRow]:
    """Эвристика без шапки — та же логика что в PDF."""
    cells_str = [str(c or "").strip() for c in cells]
    from .pdf_parser import _row_from_cells
    return _row_from_cells(cells_str, row_num)


def parse_excel(path: str) -> ParseResult:
    result = ParseResult(source_file=os.path.basename(path))
    row_num = 1

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        mapping = _detect_header(ws)
        header_found = mapping is not None
        data_start = 1

        if header_found:
            # Определяем строку с заголовком, начинаем с +1
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                row_str = " ".join(str(c or "").lower() for c in row)
                if any(kw in row_str for kw in ["сечение", "длина", "марка", "section"]):
                    data_start = i + 1
                    break

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            cells = list(row)
            if all(c is None or str(c).strip() == "" for c in cells):
                continue
            if header_found:
                r = _row_from_mapping(cells, mapping, row_num)
            else:
                r = _row_heuristic(cells, row_num)
            if r:
                result.rows.append(r)
                row_num += 1
            else:
                result.skipped_count += 1

    wb.close()
    result.total_pages = len(wb.sheetnames)
    result.parsed_count = len(result.rows)
    return result
