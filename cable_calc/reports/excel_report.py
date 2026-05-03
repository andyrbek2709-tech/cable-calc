"""Генератор Excel-отчёта по результатам расчёта."""
import io
from typing import List
from dataclasses import asdict

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# Цвета
C_HEADER_BG  = "1F3864"   # тёмно-синий
C_HEADER_FG  = "FFFFFF"
C_OK_BG      = "E2EFDA"   # светло-зелёный
C_ERR_BG     = "FCE4D6"   # светло-красный
C_WARN_BG    = "FFEB9C"   # жёлтый
C_ALT_BG     = "F2F2F2"   # серый для чередования
C_BLUE_FG    = "0000FF"   # hardcode inputs
C_BLACK_FG   = "000000"

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("№",           5),
    ("Обозначение", 18),
    ("Фазы",        6),
    ("Марка",       12),
    ("Сечение, мм²",10),
    ("Длина, м",    10),
    ("Метод",       7),
    ("Iр, А",       9),
    ("Iдоп, А",     9),
    ("ΔU, %",       8),
    ("CB, А",       8),
    ("Ткабель",     8),
    ("k_t",         7),
    ("k_gr",        7),
    ("IКЗ1ф, А",    10),
    ("IКЗ3ф, А",    10),
    ("Статус",      10),
    ("Подсказки",   40),
]


def _hdr_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=C_HEADER_FG, name="Arial", size=9)
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c


def _data_cell(ws, row, col, value, bg=None, fg=C_BLACK_FG, num_fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=9, color=fg, bold=bold)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
    if num_fmt:
        c.number_format = num_fmt
    return c


def build_excel_report(results: list, inputs: list = None) -> bytes:
    """
    results: список CableResult (dataclass) или dict
    inputs:  список CableInput (опционально, для методологии)
    Возвращает bytes xlsx-файла.
    """
    wb = openpyxl.Workbook()

    # ── Лист 1: Сводная таблица ──────────────────────────────────────
    ws = wb.active
    ws.title = "Расчёт кабелей"
    ws.freeze_panes = "A3"

    # Строка 1: заголовок документа
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = "Расчёт кабелей до 1кВ · МЭК 60364-5-52"
    title_cell.font = Font(bold=True, size=12, name="Arial", color=C_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Строка 2: заголовки колонок
    for ci, (hdr, width) in enumerate(COLUMNS, 1):
        _hdr_cell(ws, 2, ci, hdr)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 30

    # Данные
    for ri, res in enumerate(results, 3):
        r = res if isinstance(res, dict) else asdict(res)
        status = r.get("status", "?")
        bg = C_OK_BG if status == "OK" else (C_ERR_BG if status == "ERROR" else C_WARN_BG)
        alt = C_ALT_BG if ri % 2 == 0 else None
        row_bg = bg  # статус перекрывает чередование

        vals = [
            ri - 2,
            r.get("line_id", ""),
            r.get("phases", 3),
            r.get("cable_mark", ""),
            r.get("section_mm2", 0),
            r.get("length_m", 0),
            r.get("method", ""),
            r.get("i_calc_a", 0),
            r.get("i_allowable_a", 0),
            r.get("delta_u_pct", 0),
            r.get("cb_rating_a", 0),
            r.get("ambient_temp_c", 30),
            r.get("k_temp", 1),
            r.get("k_group", 1),
            r.get("i_kz_1ph_a", 0),
            r.get("i_kz_3ph_a", 0),
            status,
            "; ".join(r.get("hints", [])),
        ]
        num_fmts = [None,None,None,None,"0.0",
                    "0",None,"0.0","0.0","0.000",
                    "0","0","0.000","0.000","0","0",None,None]
        for ci, (v, nf) in enumerate(zip(vals, num_fmts), 1):
            fg = C_BLACK_FG
            if ci in (8, 9):   # токи — синий (hardcode input)
                fg = C_BLUE_FG
            _data_cell(ws, ri, ci, v, bg=row_bg, fg=fg, num_fmt=nf)

    # Итоговая строка
    last = len(results) + 3
    ws.merge_cells(f"A{last}:P{last}")
    total_ok = sum(1 for r in results
                   if (r if isinstance(r, dict) else asdict(r)).get("status") == "OK")
    c = ws.cell(row=last, column=1,
                value=f"Итого: {len(results)} кабелей · OK: {total_ok} · "
                      f"Ошибок: {len(results)-total_ok}")
    c.font = Font(bold=True, name="Arial", size=9)
    c.fill = PatternFill("solid", fgColor="D9D9D9")
    c.alignment = Alignment(horizontal="left")

    # ── Лист 2: Методология ──────────────────────────────────────────
    wm = wb.create_sheet("Методология")
    wm.column_dimensions["A"].width = 25
    wm.column_dimensions["B"].width = 70

    mrow = 1
    wm.merge_cells("A1:B1")
    wm["A1"].value = "Методология расчёта · МЭК 60364-5-52"
    wm["A1"].font = Font(bold=True, size=11, name="Arial", color=C_HEADER_FG)
    wm["A1"].fill = PatternFill("solid", fgColor=C_HEADER_BG)
    wm["A1"].alignment = Alignment(horizontal="center")
    mrow += 1

    # Методология первого результата как пример
    if results:
        r = results[0] if isinstance(results[0], dict) else asdict(results[0])
        method_data = r.get("methodology", {})
        for key, val in method_data.items():
            c_key = wm.cell(row=mrow, column=1, value=str(key))
            c_key.font = Font(bold=True, name="Arial", size=9)
            c_key.fill = PatternFill("solid", fgColor="E2EFDA")
            c_key.border = BORDER
            c_val = wm.cell(row=mrow, column=2, value=str(val))
            c_val.font = Font(name="Arial", size=9)
            c_val.border = BORDER
            c_val.alignment = Alignment(wrap_text=True)
            mrow += 1

    # Сохраняем
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
